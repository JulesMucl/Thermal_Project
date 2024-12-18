

"""
LELME2150 - Thermal cycles
Homework 3 - Steam turbine

Test code for your function

@author: Antoine Laterre
@date: October 30, 2022
"""

#
#===IMPORT PACKAGES============================================================
#
import matplotlib.pyplot as plt
import pandas as pd
import os
import sys
import numpy as np
from openpyxl import load_workbook
from openpyxl.drawing.image import Image

from Nouvelle_resolution_opti_T_surch import ORC

# region run opti_T_surch
# Définir le répertoire de sortie (vous pouvez le personnaliser)
#output_directory = "Graphs_R245fa_basic"  # Dossier pour sauvegarder les fichiers de sortie

# Charger le fichier Excel
excel_path = os.path.join(os.path.dirname(__file__), '..', 'Param', 'Projet_contraintes.xlsx')
data = pd.read_excel(excel_path)

# Créer le dictionnaire à partir des données Excel
params_excel = {row['Name']: row['Value'] for _, row in data.iterrows()}

params = {}
params.update(params_excel)
inputs = 0

my_ORC = ORC(inputs,params,display=True)

my_ORC.evaluate()

print('Pe =', my_ORC.Pe)
print('m_tot =', my_ORC.m_tot)
print('eta_cyclex =', my_ORC.eta_cyclex)
print('eta_cyclen = ', my_ORC.eta_cyclen)

#endregion run opti_T_surch




# region BON 


# # Définir le répertoire de sortie (vous pouvez le personnaliser)
# output_directory = "Graphs_R245fa_basic"  # Dossier pour sauvegarder les fichiers de sortie

# # Charger le fichier Excel
# excel_path = os.path.join(os.path.dirname(__file__), '..', 'Param', 'Projet_contraintes.xlsx')
# data = pd.read_excel(excel_path)

# # Créer le dictionnaire à partir des données Excel
# params_excel = {row['Name']: row['Value'] for _, row in data.iterrows()}

# params = {}
# params.update(params_excel)
# inputs = 0

# my_ORC = ORC(inputs,params,display=True, output_dir=output_directory)

# my_ORC.evaluate()

# print('Pe =', my_ORC.Pe)
# print('m_tot =', my_ORC.m_tot)
# print('eta_cyclex =', my_ORC.eta_cyclex)
# print('eta_cyclen = ', my_ORC.eta_cyclen)

# endregion BON


# region T3
# # Charger le fichier Excel
# excel_path = os.path.join(os.path.dirname(__file__), '..', 'Param', 'Projet_contraintes.xlsx')
# data = pd.read_excel(excel_path)

# # Créer le dictionnaire à partir des données Excel
# params_excel = {row['Name']: row['Value'] for _, row in data.iterrows()}

# params = {}
# params.update(params_excel)
# inputs = 0

# # Initialisation des variables pour T_11
# i_values = np.linspace(5, 30, 10)  # Exemple : linspace de 5°C à 30°C avec 10 points
# T4_surchauffe_values = np.linspace(5, 30, 10)  # Exemple : linspace de 5°C à 30°C avec 10 points
# results = []

# # Boucle sur T_11
# for i in i_values:
#     print(f"Running ORC model for T_11 = {i} °C")
    
#     # Mettre à jour le paramètre dans le dictionnaire
#     params['T_11'] = i  # Mise à jour dynamique
#     params['T_surchauffe_4'] = i  # Mise à jour dynamique
    
#     # Initialiser et évaluer le modèle
#     my_ORC = ORC(inputs, params, display=False)
#     my_ORC.evaluate()
    
#     # Stocker les résultats (vous pouvez choisir ce que vous voulez stocker)
#     results.append({
#         'T_11': i,
#         'eta_cycle': my_ORC.eta_cyclex,  # Rendement exergétique
#         'Pe': my_ORC.Pe,  # Puissance électrique
#         'm_tot': my_ORC.m_tot  # Débit massique total
#     })

# # Convertir les résultats en DataFrame pour analyse
# results_df = pd.DataFrame(results)

# # Afficher les résultats sous forme de tableau
# print(results_df)

# # Graphique des résultats
# plt.figure(figsize=(10, 6))
# plt.plot(results_df['T_11'], results_df['eta_cycle'], label="Rendement exergétique η", marker='o')
# #plt.plot(results_df['T_11'], results_df['Pe'], label="Puissance électrique (kW)", marker='x')
# plt.xlabel("T_11 (°C)")
# plt.ylabel("Valeurs")
# plt.title("Impact de T_11 sur le modèle ORC")
# plt.legend()
# plt.grid()


# # Sauvegarder le graphe final
# output_dir = "graphs"  # Dossier pour sauvegarder le fichier
# os.makedirs(output_dir, exist_ok=True)  # Créer le dossier s'il n'existe pas
# final_graph_filename = os.path.join(output_dir, "Rendement_energétique_T_11.png")
# plt.savefig(final_graph_filename, dpi=300)
# print(f"Graphique final sauvegardé sous : {final_graph_filename}")



# plt.show()

#  endregion T3

#region COPS
# import os
# import pandas as pd
# import numpy as np
# import matplotlib.pyplot as plt
# from openpyxl import load_workbook
# from openpyxl.drawing.image import Image

# from Nouvelle_resolution_sans_opti import ORC  # Assurez-vous d'importer correctement votre module ORC

# # Charger le fichier Excel
# excel_path = os.path.join(os.path.dirname(__file__), '..', 'Param', 'Projet_contraintes.xlsx')
# data = pd.read_excel(excel_path)

# # Créer le dictionnaire à partir des données Excel
# params_excel = {row['Name']: row['Value'] for _, row in data.iterrows()}

# params = {}
# params.update(params_excel)
# inputs = 0

# # Initialisation des variables pour T_11
# variable_1 = np.linspace(274.15, 295.13, 10)  # Plage de valeurs pour T_11
# results = []

# # Boucle sur T_11
# for i in variable_1:
#     print(f"Running ORC model for T_11 = {i} °C")
    
#     # Mettre à jour le paramètre dans le dictionnaire
#     params['T_11'] = i
    
#     # Initialiser et évaluer le modèle
#     my_ORC = ORC(inputs, params, display=False)
#     my_ORC.evaluate()
    
#     # Stocker les résultats en prenant la valeur absolue de loss_conden et loss_condex
#     results.append({
#         'T_11': i,
#         'eta_cyclex': my_ORC.eta_cyclex,        # Rendement exergétique
#         'eta_cyclen': my_ORC.eta_cyclen,        # Rendement thermique
#         'Pe': my_ORC.Pe,                        # Puissance électrique
#         'm_tot': my_ORC.m_tot,                  # Débit massique total
#         'loss_conden': abs(my_ORC.loss_conden), # Perte au condenseur
#         'loss_condex': abs(my_ORC.loss_condex)  # Perte exergétique au condenseur
#     })

# # Convertir les résultats en DataFrame
# results_df = pd.DataFrame(results)

# # Créer un dossier pour les graphiques
# output_dir = "Graphs_R245ca_COPS"
# os.makedirs(output_dir, exist_ok=True)

# # Sauvegarder le tableau complet dans un fichier Excel
# output_excel_path = os.path.join(output_dir, "Resultats_ORC_T_11.xlsx")
# results_df.to_excel(output_excel_path, index=False, sheet_name="Tableau")

# # Graphique 1 : Rendement exergétique η_cyclex
# plt.figure(figsize=(8, 6))
# plt.plot(results_df['T_11'], results_df['eta_cyclex'], marker='o', label="Rendement exergétique η_cyclex")
# plt.xlabel("T_11 (°C)")
# plt.ylabel("η_cyclex (%)")
# plt.title("Impact de T_11 sur le Rendement exergétique")
# plt.legend()
# plt.grid()
# graph_1_filename = os.path.join(output_dir, "Graph_eta_cyclex.png")
# plt.savefig(graph_1_filename, dpi=300)
# plt.show()

# # Graphique 2 : Rendement thermique η_cyclen
# plt.figure(figsize=(8, 6))
# plt.plot(results_df['T_11'], results_df['eta_cyclen'], marker='o', color='orange', label="Rendement thermique η_cyclen")
# plt.xlabel("T_11 (°C)")
# plt.ylabel("η_cyclen (%)")
# plt.title("Impact de T_11 sur le Rendement thermique")
# plt.legend()
# plt.grid()
# graph_6_filename = os.path.join(output_dir, "Graph_eta_cyclen.png")
# plt.savefig(graph_6_filename, dpi=300)
# plt.show()

# # Graphique 3 : Puissance électrique Pe
# plt.figure(figsize=(8, 6))
# plt.plot(results_df['T_11'], results_df['Pe'], marker='x', color='r', label="Puissance électrique (kW)")
# plt.xlabel("T_11 (°C)")
# plt.ylabel("Pe (kW)")
# plt.title("Impact de T_11 sur la Puissance électrique")
# plt.legend()
# plt.grid()
# graph_2_filename = os.path.join(output_dir, "Graph_Pe.png")
# plt.savefig(graph_2_filename, dpi=300)
# plt.show()

# # Graphique 4 : Débit massique total m_tot
# plt.figure(figsize=(8, 6))
# plt.plot(results_df['T_11'], results_df['m_tot'], marker='s', color='g', label="Débit massique total (kg/s)")
# plt.xlabel("T_11 (°C)")
# plt.ylabel("m_tot (kg/s)")
# plt.title("Impact de T_11 sur le Débit massique total")
# plt.legend()
# plt.grid()
# graph_3_filename = os.path.join(output_dir, "Graph_m_tot.png")
# plt.savefig(graph_3_filename, dpi=300)
# plt.show()

# # Graphique 5 : Perte au condenseur loss_conden
# plt.figure(figsize=(8, 6))
# plt.plot(results_df['T_11'], results_df['loss_conden'], marker='^', color='b', label="Perte au condenseur (kW)")
# plt.xlabel("T_11 (°C)")
# plt.ylabel("Loss Conden (kW)")
# plt.title("Pertes énergétiques au condenseur")
# plt.legend()
# plt.grid()
# graph_4_filename = os.path.join(output_dir, "Graph_loss_conden.png")
# plt.savefig(graph_4_filename, dpi=300)
# plt.show()

# # Graphique 6 : Nouvelle perte condex
# plt.figure(figsize=(8, 6))
# plt.plot(results_df['T_11'], results_df['loss_condex'], marker='v', color='m', label="Perte condenseur externe (kW)")
# plt.xlabel("T_11 (°C)")
# plt.ylabel("Loss Condex (kW)")
# plt.title("Pertes exergétiques au condenseur")
# plt.legend()
# plt.grid()
# graph_5_filename = os.path.join(output_dir, "Graph_loss_condex.png")
# plt.savefig(graph_5_filename, dpi=300)
# plt.show()

# # Insérer les graphiques dans le fichier Excel
# wb = load_workbook(output_excel_path)
# ws = wb.create_sheet("Graphiques")

# # Insérer chaque graphique dans des cellules différentes
# img1 = Image(graph_1_filename)
# img2 = Image(graph_2_filename)
# img3 = Image(graph_3_filename)
# img4 = Image(graph_4_filename)
# img5 = Image(graph_5_filename)
# img6 = Image(graph_6_filename)  # Nouvelle image pour eta_cyclen

# ws.add_image(img1, "A1")
# ws.add_image(img6, "G1")  # Position du graphique eta_cyclen
# ws.add_image(img2, "A20")
# ws.add_image(img3, "G20")
# ws.add_image(img4, "A40")
# ws.add_image(img5, "G40")

# wb.save(output_excel_path)

# print(f"Fichier Excel sauvegardé avec graphiques sous : {output_excel_path}")


# endregion COPS

























# eta_en = my_ORC.eta_toten
# p1 = my_ORC.p1_guess_plot
# p2 = my_ORC.p2_guess_plot
# p5 = my_ORC.p5_guess_plot



# plt.figure(figsize=(10, 6))
# plt.plot(p1, label="p1_guess_plot", linestyle='-', marker='o')
# plt.plot(p2, label="p2_guess_plot", linestyle='--', marker='x')
# plt.plot(p5, label="p5_guess_plot", linestyle='-.', marker='s')

# # Ajout des légendes et titres
# plt.title("Comparison of p1_guess_plot, p2_guess_plot, and p5_guess_plot")
# plt.xlabel("Index")
# plt.ylabel("Values")
# plt.legend()
# plt.grid(True)

# # Affichage du graphe
# plt.show()



















# """
# LELME2150 - Thermal cycles
# Homework 3 - Steam turbine

# Test code for your function

# @author: Antoine Laterre
# @date: October 30, 2022
# """

# #
# #===IMPORT PACKAGES============================================================
# #
# import matplotlib.pyplot as plt
# import pandas as pd
# import os
# import sys

# from Nouvelle_resolution_sans_opti import ORC

# # Charger le fichier Excel
# excel_path = os.path.join(os.path.dirname(__file__), '..', 'Param', 'Projet_contraintes.xlsx')
# data = pd.read_excel(excel_path)

# # Créer le dictionnaire à partir des données Excel
# params_excel = {row['Name']: row['Value'] for _, row in data.iterrows()}

# params = {}
# params.update(params_excel)
# inputs = 0



# my_ORC = ORC(inputs,params,False)
# try:
#     my_ORC.evaluate()
# except ValueError as e:
#     print(f"Une erreur est survenue dans my_ORC.evaluate(): {e}")

